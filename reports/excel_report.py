import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def export_to_excel(results: list, output_path: str):
    """
    將選股結果輸出成 Excel。
    """

    df = pd.DataFrame(results)

    if df.empty:
        print("沒有資料可輸出")
        return

    df = df.sort_values(by="分數", ascending=False)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="選股結果")

        ws = writer.book["選股結果"]

        # 凍結首列
        ws.freeze_panes = "A2"

        # 加上篩選器
        ws.auto_filter.ref = ws.dimensions

        # 標題格式
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 找欄位位置
        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            headers[cell.value] = idx

        signal_col = headers.get("訊號")
        flag_col = headers.get("風險標註")
        reason_col = headers.get("系統解讀")

        # 依照訊號上色
        if signal_col:
            for row in range(2, ws.max_row + 1):
                signal = ws.cell(row=row, column=signal_col).value

                if signal == "買進觀察":
                    fill = PatternFill("solid", fgColor="C6EFCE")
                elif signal == "留意追蹤":
                    fill = PatternFill("solid", fgColor="FFEB9C")
                elif signal == "高風險排除":
                    fill = PatternFill("solid", fgColor="FFC7CE")
                elif signal == "有陷阱":
                    fill = PatternFill("solid", fgColor="F4B084")
                else:
                    fill = None

                if fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill

        # 風險標註欄位加強顯示
        if flag_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=flag_col)
                if cell.value:
                    cell.font = Font(color="9C0006", bold=True)

        # 系統解讀換行
        if reason_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=reason_col).alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        # 全表基本對齊
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True if cell.column in [reason_col, flag_col] else False
                )

        # 自動欄寬
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            width = min(max_length + 4, 45)
            ws.column_dimensions[column_letter].width = width

        # 特別調整系統解讀欄、風險欄寬度
        if reason_col:
            ws.column_dimensions[get_column_letter(reason_col)].width = 55

        if flag_col:
            ws.column_dimensions[get_column_letter(flag_col)].width = 35
